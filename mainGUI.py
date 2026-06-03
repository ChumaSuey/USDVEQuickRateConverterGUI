import csv
import datetime  # Añadido para automatizar las fechas por defecto
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# Importación de la capa de datos y formato desde el backend modular
from backend import format_currency_ve, fetch_all_bcv_history

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class BcvDashboardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BCV Dollar Tracker - Contabilidad Profesional")
        
        # Resolución adaptativa para pantallas pequeñas (Responsive window sizing)
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        
        app_width = min(710, int(screen_width * 0.85))
        app_height = min(700, int(screen_height * 0.85))
        
        # Centrar la ventana dinámicamente
        x = int((screen_width / 2) - (app_width / 2))
        y = int((screen_height / 2) - (app_height / 2))
        
        self.root.geometry(f"{app_width}x{app_height}+{x}+{y}")
        self.root.minsize(600, 450)  # Evita que se encoja demasiado y rompa la UI
        self.root.configure(bg="#1e1e2e")
        
        self.all_historical_records = []
        self.current_view_records = []
        self.is_filtered_view = False
        self.selected_month_year_str = ""
        self.current_currency = "dolares"

        self.months_map = {
            "Enero": "01", "Febrero": "02", "Marzo": "03", "Abril": "04",
            "Mayo": "05", "Junio": "06", "Julio": "07", "Agosto": "08",
            "Septiembre": "09", "Octubre": "10", "Noviembre": "11", "Diciembre": "12"
        }
        
        self.setup_styles()
        self.create_header()
        self.create_metrics_grid()
        self.create_filter_bar()
        self.create_data_table()
        self.create_footer_controls()
        
        self.load_api_data_first_time()

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", 
                        background="#252538", fieldbackground="#252538", foreground="#cdd6f4", 
                        rowheight=28, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", 
                        background="#11111b", foreground="#cdd6f4", font=("Segoe UI", 10, "bold"), borderwidth=0)
        style.map("Treeview", background=[("selected", "#45475a")], foreground=[("selected", "#a6e3a1")])
        style.configure("TCombobox", fieldbackground="#1a2744", background="#11111b", foreground="#cdd6f4")
        style.map("TCombobox",
                  fieldbackground=[("readonly", "#1a2744")],
                  foreground=[("readonly", "#cdd6f4")],
                  selectbackground=[("readonly", "#3b82f6"), ("readonly", "focus", "#3b82f6")],
                  selectforeground=[("readonly", "#ffffff"), ("readonly", "focus", "#ffffff")])
        
        self.root.option_add("*TCombobox*Listbox.background", "#1a2744")
        self.root.option_add("*TCombobox*Listbox.foreground", "#cdd6f4")
        self.root.option_add("*TCombobox*Listbox.selectBackground", "#3b82f6")
        self.root.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")

    def create_header(self):
        header_frame = tk.Frame(self.root, bg="#11111b", height=80)
        header_frame.pack(fill="x", padx=15, pady=(15, 5))
        
        self.lbl_main_rate = tk.Label(
            header_frame, text="Conectando al BCV...", font=("Segoe UI", 22, "bold"), fg="#a6e3a1", bg="#11111b"
        )
        self.lbl_main_rate.pack(side="left", padx=20, pady=15)
        
        self.lbl_status_date = tk.Label(
            header_frame, text="Procesando...", font=("Segoe UI", 10, "italic"), fg="#a6adc8", bg="#11111b"
        )
        self.lbl_status_date.pack(side="right", padx=20, pady=20)

    def create_metrics_grid(self):
        metrics_frame = tk.Frame(self.root, bg="#1e1e2e")
        metrics_frame.pack(fill="x", padx=15, pady=5)
        
        for i in range(3): metrics_frame.columnconfigure(i, weight=1)
        
        c1 = tk.Frame(metrics_frame, bg="#252538")
        c1.grid(row=0, column=0, padx=5, sticky="nsew")
        tk.Label(c1, text="MÍNIMO PERIODO", font=("Segoe UI", 9, "bold"), fg="#89b4fa", bg="#252538").pack(pady=(10,2))
        self.lbl_metric_low = tk.Label(c1, text="--,--", font=("Segoe UI", 14, "bold"), fg="#cdd6f4", bg="#252538")
        self.lbl_metric_low.pack(pady=(0,10))

        c2 = tk.Frame(metrics_frame, bg="#252538")
        c2.grid(row=0, column=1, padx=5, sticky="nsew")
        tk.Label(c2, text="MÁXIMO PERIODO", font=("Segoe UI", 9, "bold"), fg="#f38ba8", bg="#252538").pack(pady=(10,2))
        self.lbl_metric_high = tk.Label(c2, text="--,--", font=("Segoe UI", 14, "bold"), fg="#cdd6f4", bg="#252538")
        self.lbl_metric_high.pack(pady=(0,10))

        c3 = tk.Frame(metrics_frame, bg="#252538")
        c3.grid(row=0, column=2, padx=5, sticky="nsew")
        tk.Label(c3, text="VARIACIÓN PERIODO", font=("Segoe UI", 9, "bold"), fg="#f9e2af", bg="#252538").pack(pady=(10,2))
        self.lbl_metric_var = tk.Label(c3, text="0,00%", font=("Segoe UI", 14, "bold"), fg="#cdd6f4", bg="#252538")
        self.lbl_metric_var.pack(pady=(0,10))

    def create_filter_bar(self):
        filter_frame = tk.LabelFrame(self.root, text=" 🔍 Filtros y Calendario Continuo ", font=("Segoe UI", 9, "bold"), fg="#bac2de", bg="#1e1e2e", bd=1, relief="groove")
        filter_frame.pack(fill="x", padx=15, pady=10, ipady=5)
        
        tk.Label(filter_frame, text="Moneda:", font=("Segoe UI", 10), fg="#cdd6f4", bg="#1e1e2e").pack(side="left", padx=(5, 2))
        self.combo_currency = ttk.Combobox(filter_frame, values=["USD - Dólar", "EUR - Euro"], width=11, state="readonly")
        self.combo_currency.pack(side="left", padx=(0, 5))
        self.combo_currency.set("USD - Dólar")
        self.combo_currency.bind("<<ComboboxSelected>>", self.on_currency_change)
        
        tk.Label(filter_frame, text="Mes:", font=("Segoe UI", 10), fg="#cdd6f4", bg="#1e1e2e").pack(side="left", padx=(5, 2))
        months_list = list(self.months_map.keys())
        self.combo_month = ttk.Combobox(filter_frame, values=months_list, width=10, state="readonly")
        self.combo_month.pack(side="left", padx=2)
        
        tk.Label(filter_frame, text="Año:", font=("Segoe UI", 10), fg="#cdd6f4", bg="#1e1e2e").pack(side="left", padx=(5, 2))
        self.combo_year = ttk.Combobox(filter_frame, values=["2024", "2025", "2026"], width=6, state="readonly")
        self.combo_year.pack(side="left", padx=2)
        
        # --- ASIGNACIÓN DINÁMICA DE FECHA ACTUAL ---
        now = datetime.datetime.now()
        self.combo_month.set(months_list[now.month - 1])  # Determina el mes actual en base a índice 0
        self.combo_year.set(str(now.year))               # Determina el año actual
        
        btn_filter = tk.Button(
            filter_frame, text="Filtrar", font=("Segoe UI", 9, "bold"),
            bg="#89b4fa", fg="#11111b", activebackground="#b4befe", bd=0, padx=8, cursor="hand2",
            command=self.apply_monthly_filter
        )
        btn_filter.pack(side="left", padx=8)
        
        btn_reset = tk.Button(
            filter_frame, text="Ver Últimos 30 Días", font=("Segoe UI", 9),
            bg="#45475a", fg="#cdd6f4", activebackground="#585b70", bd=0, padx=8, cursor="hand2",
            command=self.reset_to_default_view
        )
        btn_reset.pack(side="right", padx=5)

    def create_data_table(self):
        table_frame = tk.Frame(self.root, bg="#1e1e2e")
        table_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        lbl_hint = tk.Label(
            table_frame, text="💡 Las filas tenues indican fines de semana o feriados (tasa arrastrada legalmente).",
            font=("Segoe UI", 9, "italic"), fg="#a6adc8", bg="#1e1e2e"
        )
        lbl_hint.pack(anchor="w", pady=(0, 5))
        
        self.tree = ttk.Treeview(table_frame, columns=("Fecha", "Tasa"), show="headings")
        self.tree.heading("Fecha", text="FECHA CALENDARIO")
        self.tree.heading("Tasa", text="TASA OFICIAL BCV (Bs.)")
        
        self.tree.column("Fecha", anchor="center", width=220)
        self.tree.column("Tasa", anchor="center", width=410)
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.tree.tag_configure("filled", foreground="#6c7086")
        self.tree.bind("<Double-1>", self.copy_rate_to_clipboard)

    def create_footer_controls(self):
        """Genera los controles inferiores incluyendo el selector directo de formato en el GUI."""
        footer = tk.Frame(self.root, bg="#1e1e2e")
        footer.pack(fill="x", padx=15, pady=15)
        
        btn_refresh = tk.Button(
            footer, text="🔄 Recargar API", font=("Segoe UI", 10, "bold"),
            bg="#313244", fg="#cdd6f4", bd=0, padx=15, pady=8, cursor="hand2",
            command=self.load_api_data_first_time
        )
        btn_refresh.pack(side="left")
        
        # --- OPCIONES DE CONFIGURACIÓN DIRECTA EN EL PANEL ---
        btn_export = tk.Button(
            footer, text="📊 Exportar Reporte", font=("Segoe UI", 10, "bold"),
            bg="#a6e3a1", fg="#11111b", activebackground="#94e2d5", bd=0, padx=20, pady=8, cursor="hand2",
            command=self.export_data
        )
        btn_export.pack(side="right")

        self.combo_format = ttk.Combobox(footer, values=["Excel (.xlsx)", "CSV (.csv)"], width=13, state="readonly")
        self.combo_format.pack(side="right", padx=10, ipady=3)
        
        # Validación en tiempo de diseño si openpyxl está presente en el entorno ejecutor
        if OPENPYXL_AVAILABLE:
            self.combo_format.set("Excel (.xlsx)")
        else:
            self.combo_format.set("CSV (.csv)")
            self.combo_format.config(state="disabled") # Bloqueo seguro preventiva contra fallos

        tk.Label(footer, text="Formato:", font=("Segoe UI", 10), fg="#cdd6f4", bg="#1e1e2e").pack(side="right")

    def copy_rate_to_clipboard(self, event):
        selected_item = self.tree.selection()
        if not selected_item: return
            
        row_values = self.tree.item(selected_item[0], "values")
        clean_rate = row_values[1].split()[0]
        
        self.root.clipboard_clear()
        self.root.clipboard_append(clean_rate)
        
        old_text = self.lbl_status_date.cget("text")
        self.lbl_status_date.config(text=f"✅ ¡{clean_rate} Copiado!", fg="#a6e3a1")
        self.root.after(2000, lambda: self.lbl_status_date.config(text=old_text, fg="#a6adc8"))

    def on_currency_change(self, event=None):
        selection = self.combo_currency.get()
        if "EUR" in selection:
            self.current_currency = "euros"
        else:
            self.current_currency = "dolares"
        self.load_api_data_first_time()

    def load_api_data_first_time(self):
        self.lbl_status_date.config(text="Procesando Base de Datos...", fg="#a6adc8")
        self.all_historical_records = fetch_all_bcv_history(currency=self.current_currency)
        
        if not self.all_historical_records:
            messagebox.showerror("Sin Conexión a Internet", "No se pudo conectar con el servidor de tasas.\n\nPor favor, verifica la red.")
            self.lbl_main_rate.config(text="Desconectado (Sin Red)", fg="#f38ba8")
            self.lbl_status_date.config(text="Conexión fallida.", fg="#f38ba8")
            return
            
        self.reset_to_default_view()

    def reset_to_default_view(self):
        self.is_filtered_view = False
        self.current_view_records = self.all_historical_records[-30:]
        
        live_baseline = [r for r in self.all_historical_records if not r["is_filled"]][-1]
        currency_sym = "USD" if getattr(self, "current_currency", "dolares") == "dolares" else "EUR"
        self.lbl_main_rate.config(text=f"1 {currency_sym} = {format_currency_ve(live_baseline['rate'])} Bs.")
        self.lbl_status_date.config(text=f"Últimos 30 Días Corridos (Tasa Activa: {live_baseline['date']})", fg="#a6adc8")
        
        self.render_records_to_ui()

    def apply_monthly_filter(self):
        if not self.all_historical_records:
            messagebox.showwarning("Sin datos", "No hay historial disponible.")
            return
            
        selected_month_name = self.combo_month.get()
        selected_year = self.combo_year.get()
        month_code = self.months_map[selected_month_name]
        
        target_prefix = f"{selected_year}-{month_code}"
        self.selected_month_year_str = f"{selected_year}_{month_code}"
        
        filtered = [r for r in self.all_historical_records if r["date"].startswith(target_prefix)]
        
        if not filtered:
            messagebox.showinfo("Búsqueda Vacía", f"No hay datos para: {selected_month_name} {selected_year}.")
            return
            
        self.is_filtered_view = True
        self.current_view_records = filtered
        
        self.lbl_status_date.config(text=f"Calendario Completo: {selected_month_name} {selected_year}", fg="#89b4fa")
        self.render_records_to_ui()

    def render_records_to_ui(self):
        for item in self.tree.get_children(): self.tree.delete(item)
            
        if not self.current_view_records: return
            
        rates = [r["rate"] for r in self.current_view_records]
        low_val = min(rates)
        high_val = max(rates)
        pct_change = ((rates[-1] - rates[0]) / rates[0]) * 100
        
        self.lbl_metric_low.config(text=f"{format_currency_ve(low_val)} Bs.")
        self.lbl_metric_high.config(text=f"{format_currency_ve(high_val)} Bs.")
        
        self.lbl_metric_var.config(text=f"{pct_change:+.2f}%".replace('.', ','))
        self.lbl_metric_var.config(fg="#f38ba8" if pct_change > 0 else "#a6e3a1")
        
        for r in reversed(self.current_view_records):
            if r["is_filled"]:
                self.tree.insert("", "end", values=(r["date"], f"{format_currency_ve(r['rate'])} Bs. (Feriado/FDS)"), tags=("filled",))
            else:
                self.tree.insert("", "end", values=(r["date"], f"{format_currency_ve(r['rate'])} Bs."))

    def export_data(self):
        """Lee la configuración seleccionada en el GUI para el guardado inmediato."""
        if not self.current_view_records:
            messagebox.showwarning("Sin datos", "No hay información para exportar.")
            return

        selected_format = self.combo_format.get()
        
        if "Excel" in selected_format:
            self.export_to_xlsx()
        else:
            self.export_to_csv()

    def export_to_xlsx(self):
        cur_code = "USD" if getattr(self, "current_currency", "dolares") == "dolares" else "EUR"
        default_name = f"Calendario_Continuo_BCV_{cur_code}_{self.selected_month_year_str}.xlsx" if self.is_filtered_view else f"Calendario_Continuo_BCV_{cur_code}_30_Dias.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Libro de Excel", "*.xlsx")],
            title="Guardar Calendario Continuo (Excel)",
            initialfile=default_name
        )
        if not file_path: return
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Calendario Continuo"
            ws.views.sheetView[0].showGridLines = True
            
            header_fill = PatternFill(start_color="1F385C", end_color="1F385C", fill_type="solid")
            filled_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Segoe UI", size=11)
            filled_font = Font(name="Segoe UI", size=11, italic=True, color="595959")
            
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            
            ws.append(["Fecha", "Tasa BCV (Bs.)"])
            for col_num in range(1, 3):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align
            
            for record in self.current_view_records:
                ws.append([record["date"], round(record["rate"], 2)])
                row_idx = ws.max_row
                
                cell_date = ws.cell(row=row_idx, column=1)
                cell_date.alignment = center_align; cell_date.border = thin_border
                
                cell_rate = ws.cell(row=row_idx, column=2)
                cell_rate.alignment = right_align; cell_rate.border = thin_border
                cell_rate.number_format = '#,##0.00'
                
                if record["is_filled"]:
                    cell_date.font = filled_font; cell_date.fill = filled_row_fill
                    cell_rate.font = filled_font; cell_rate.fill = filled_row_fill
                else:
                    cell_date.font = data_font
                    cell_rate.font = data_font
            
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = format_currency_ve(cell.value) if isinstance(cell.value, (int, float)) else str(cell.value)
                    if len(val_str) > max_len: max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 6, 16)
                
            wb.save(file_path)
            messagebox.showinfo("Éxito", "¡Libro de Excel creado impecablemente!")
        except Exception as e:
            messagebox.showerror("Error de guardado", f"No se pudo escribir el archivo:\n{e}")

    def export_to_csv(self):
        cur_code = "USD" if getattr(self, "current_currency", "dolares") == "dolares" else "EUR"
        default_name = f"Calendario_BCV_{cur_code}_{self.selected_month_year_str}.csv" if self.is_filtered_view else f"Calendario_BCV_{cur_code}_30_Dias.csv"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("Archivos CSV", "*.csv")], title="Guardar tasas (CSV)", initialfile=default_name
        )
        if not file_path: return
        try:
            with open(file_path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(["Fecha", "Tasa BCV (Bs.)"])
                for record in self.current_view_records:
                    writer.writerow([record["date"], format_currency_ve(record["rate"])])
            messagebox.showinfo("Éxito (CSV)", "Guardado como CSV con calendario continuo.")
        except Exception as e:
            messagebox.showerror("Error de guardado", f"No se pudo escribir el archivo:\n{e}")


if __name__ == "__main__":
    app_window = tk.Tk()
    app = BcvDashboardApp(app_window)
    app_window.mainloop()